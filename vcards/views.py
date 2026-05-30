from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.contrib.auth.hashers import check_password
from django.contrib.auth.models import User
from django.conf import settings
from django.db.models import Count, Q
from django.db.models.functions import TruncDate
from django.template import TemplateDoesNotExist
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from datetime import timedelta
from io import BytesIO
import json
import re
import textwrap
from urllib.parse import urlencode
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import qrcode
import requests
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from .models import StudentProfile, College, ProfileActivity, Skill

CONTACT_TEMPLATES = {
    'general': ['student_digital_card.html'],
    'vip': ['student_digital_card.html'],
    'premium': ['student_digital_card.html'],
}
CONTACT_TEMPLATE_META = {
    'student_digital_card.html': {
        'label': 'School Digital ID',
        'description': 'Mobile-first QR/NFC digital profile for students, teachers, and staff.',
    },
}
PORTFOLIO_TEMPLATES = {}
PRINT_TEMPLATES = {
    'front_academic_blue': {'label': 'Crest Heritage', 'description': 'Classic premium front with a stronger academic identity and portrait-led hierarchy.'},
    'front_modern_green': {'label': 'Circuit Edge', 'description': 'Modern smart-tech front that suits QR and NFC-enabled school cards.'},
    'front_maroon_crest': {'label': 'Noble Arch', 'description': 'Refined premium front with a formal institutional look.'},
    'front_prism_shift': {'label': 'Prism Shift', 'description': 'Bold geometric front with a cleaner contemporary school feel.'},
    'front_official_wave': {'label': 'Official Wave', 'description': 'Formal school front with soft waves and school branding.'},
}

PRINT_CARD_TYPES = [
    {'value': 'id_card', 'label': 'ID Card'},
]

PRINT_ORIENTATIONS = [
    {'value': 'portrait', 'label': 'Portrait'},
    {'value': 'landscape', 'label': 'Landscape'},
]

PRINT_FRONT_THEMES = {
    'front_academic_blue': {
        'label': 'Crest Heritage',
        'description': 'Timeless academic front inspired by formal institutional credentials.',
        'palette': ['#1a3a5c', '#f7f5f0', '#1a1a2a', '#c9a84c'],
    },
    'front_modern_green': {
        'label': 'Circuit Edge',
        'description': 'Dark modern smart-tech front with stronger QR and NFC energy.',
        'palette': ['#00d4aa', '#12121e', '#e0e0f0', '#00856a'],
    },
    'front_maroon_crest': {
        'label': 'Noble Arch',
        'description': 'Refined premium front with a softer formal school character.',
        'palette': ['#2d4a22', '#faf8f4', '#222222', '#d4a853'],
    },
    'front_prism_shift': {
        'label': 'Prism Shift',
        'description': 'Contemporary diagonal layout that still feels school-approved and print practical.',
        'palette': ['#c43e3e', '#ffffff', '#1a1a1a', '#962e2e'],
    },
    'front_official_wave': {
        'label': 'Official Wave',
        'description': 'Formal school front with soft waves, circular portrait, and a document-style signature overlay.',
        'palette': ['#1e3a8a', '#ffffff', '#1f2937', '#f97316'],
    },
}

PRINT_BACK_THEMES = {
    'back_qr_clean': {
        'label': 'Crest Heritage Back',
        'description': 'Classic premium back with QR, signature, and school contact details.',
        'palette': ['#1a3a5c', '#f7f5f0', '#1a1a2a', '#c9a84c'],
    },
    'back_contact_grid': {
        'label': 'Circuit Edge Back',
        'description': 'Modern back with QR as the hero and clearer digital-verification language.',
        'palette': ['#00d4aa', '#12121e', '#e0e0f0', '#00856a'],
    },
    'back_secure_band': {
        'label': 'Noble Arch Back',
        'description': 'Elegant back with QR, principal sign, and formal support details.',
        'palette': ['#2d4a22', '#faf8f4', '#222222', '#d4a853'],
    },
    'back_prism_shift': {
        'label': 'Prism Shift Back',
        'description': 'Bold geometric back with QR, NFC, and signature zones.',
        'palette': ['#c43e3e', '#ffffff', '#1a1a1a', '#962e2e'],
    },
    'back_official_wave': {
        'label': 'Official Wave Back',
        'description': 'Formal back with QR, school contact, and clear terms-style note section.',
        'palette': ['#1e3a8a', '#ffffff', '#1f2937', '#f97316'],
    },
}

STUDIO_FRONT_THEME_KEYS = tuple(PRINT_FRONT_THEMES.keys())
STUDIO_BACK_THEME_KEYS = tuple(PRINT_BACK_THEMES.keys())

LEGACY_PRINT_TEMPLATE_MAP = {
    'print_classic.html': 'front_academic_blue',
    'print_school.html': 'front_maroon_crest',
    'print_modern.html': 'front_modern_green',
}

SOCIAL_CHOICES = [
    "linkedin", "instagram", "facebook", "messenger", "whatsapp",
    "twitter", "youtube", "tiktok", "github", "figma", "upwork"
]

SCHOOL_ROLE_CHOICES = [
    'Teacher',
    'Senior Teacher',
    'Lecturer',
    'Professor',
    'Assistant Professor',
    'Head of Department',
    'Coordinator',
    'Vice Principal',
    'Principal',
    'Chairman',
    'Dean',
    'Faculty Mentor',
]

FACULTY_ROLE_KEYWORDS = (
    'teacher', 'faculty', 'lecturer', 'professor', 'principal', 'chairman',
    'chairperson', 'hod', 'head', 'coordinator', 'dean', 'mentor', 'instructor',
    'trainer', 'academic'
)

ACADEMIC_LEVEL_CHOICES = StudentProfile.ACADEMIC_LEVEL_CHOICES
GENDER_CHOICES = [choice for choice in StudentProfile.GENDER_CHOICES if choice[0]]


def _student_edit_session_key(student_id):
    return f'can_edit_student_{student_id}'


def _profile_supports_self_service(student):
    return bool(student and student.member_type == 'teacher')


def _is_student_edit_authorized(request, student_id):
    try:
        student = StudentProfile.objects.only('id', 'member_type').get(id=student_id)
    except StudentProfile.DoesNotExist:
        return False
    if not _profile_supports_self_service(student):
        request.session.pop(_student_edit_session_key(student_id), None)
        return False
    return request.session.get(_student_edit_session_key(student_id), False)


def _is_super_admin(user):
    return user.is_authenticated and (user.is_superuser or user.is_staff)


def _get_managed_school(user):
    if not user.is_authenticated:
        return None
    return College.objects.filter(admin_user=user).first()


def _get_owned_profile(user):
    if not user.is_authenticated:
        return None
    return StudentProfile.objects.filter(auth_user=user).first()


def _get_user_role(user):
    if not user.is_authenticated:
        return 'public'
    if _is_super_admin(user):
        return 'super_admin'
    if _get_managed_school(user):
        return 'school_admin'
    owned_profile = _get_owned_profile(user)
    if owned_profile and _profile_supports_self_service(owned_profile):
        return 'teacher'
    return 'public'


def _can_manage_school(user, college):
    if _is_super_admin(user):
        return True
    return user.is_authenticated and college.admin_user_id == user.id


def _can_manage_profile(user, student):
    if _is_super_admin(user):
        return True
    if (
        student.auth_user_id
        and student.auth_user_id == user.id
        and _profile_supports_self_service(student)
    ):
        return True
    managed_school = _get_managed_school(user)
    return bool(managed_school and student.college_id == managed_school.id)


def _require_school_access(request, college):
    if _can_manage_school(request.user, college):
        return None
    messages.error(request, 'You do not have permission to access this school workspace.')
    return redirect('dashboard_login')


def _require_profile_access(request, student):
    if _can_manage_profile(request.user, student) or _is_student_edit_authorized(request, student.id):
        return None
    messages.error(request, 'You do not have permission to access this profile.')
    return redirect('dashboard_login')


def _ensure_unique_username(base_username, exclude_user_id=None):
    username = (base_username or 'user').strip() or 'user'
    candidate = username
    suffix = 1
    while User.objects.filter(username=candidate).exclude(id=exclude_user_id).exists():
        suffix += 1
        candidate = f'{username}{suffix}'
    return candidate


def _sync_profile_auth_user(student, raw_password=None):
    if not _profile_supports_self_service(student):
        if student.auth_user_id:
            student.auth_user = None
        return None
    profile_user = student.auth_user
    target_username = _ensure_unique_username(student.username, profile_user.id if profile_user else None)
    if not profile_user:
        profile_user = User(username=target_username)
    else:
        profile_user.username = target_username
    profile_user.first_name = student.name
    profile_user.email = student.email or ''
    if raw_password:
        profile_user.set_password(raw_password)
    profile_user.save()
    if student.auth_user_id != profile_user.id or student.username != target_username:
        student.auth_user = profile_user
        student.username = target_username
    return profile_user


def _sync_school_admin_user(college, username, raw_password):
    username = _ensure_unique_username(username, college.admin_user_id)
    admin_user = college.admin_user or User(username=username)
    admin_user.username = username
    admin_user.first_name = college.name
    admin_user.email = college.email or ''
    admin_user.is_staff = False
    admin_user.is_superuser = False
    if raw_password:
        admin_user.set_password(raw_password)
    admin_user.save()
    college.admin_user = admin_user
    return admin_user


def _generate_profile_password(name):
    clean_name = ''.join(ch for ch in (name or 'user') if ch.isalpha()) or 'user'
    first_three = clean_name[:3]
    return f'{first_three.upper()}@@123{first_three.lower()}'


def _build_dashboard_query(school=None):
    if not school:
        return ''
    return f"?{urlencode({'school': school.id})}"


def _resolve_dashboard_school(request, required=False):
    if _is_super_admin(request.user):
        school_id = (request.GET.get('school') or request.POST.get('school') or '').strip()
        schools = College.objects.order_by('name')
        school = None
        if school_id:
            school = schools.filter(id=school_id).first()
        elif required:
            school = schools.first()
        return school, schools

    school = _get_managed_school(request.user)
    schools = College.objects.filter(id=school.id) if school else College.objects.none()
    return school, schools


def _school_dashboard_context(request, active_module, school=None, schools=None):
    nav_school_query = _build_dashboard_query(school)
    return {
        'active_module': active_module,
        'current_school': school,
        'school_options': schools or College.objects.none(),
        'nav_school_query': nav_school_query,
        'is_super_admin': _is_super_admin(request.user),
    }


def _school_member_queryset(school, member_type=None):
    queryset = StudentProfile.objects.select_related('college').filter(
        profile_category='school',
        college=school,
    ).order_by('name')
    if member_type:
        queryset = queryset.filter(member_type=member_type)
    return queryset


def _apply_student_filters(queryset, request):
    academic_level = (request.GET.get('academic_level') or request.POST.get('academic_level') or '').strip()
    section = (request.GET.get('section') or request.POST.get('section') or '').strip()
    search = (request.GET.get('q') or request.POST.get('q') or '').strip()

    if academic_level:
        queryset = queryset.filter(academic_level=academic_level)
    if section:
        queryset = queryset.filter(section__iexact=section)
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search)
            | Q(email__icontains=search)
            | Q(phone__icontains=search)
            | Q(roll_number__icontains=search)
        )
    return queryset, academic_level, section, search


def _apply_teacher_filters(queryset, request):
    search = (request.GET.get('q') or request.POST.get('q') or '').strip()
    role = (request.GET.get('role') or request.POST.get('role') or '').strip()

    if search:
        queryset = queryset.filter(
            Q(name__icontains=search)
            | Q(email__icontains=search)
            | Q(phone__icontains=search)
            | Q(role__icontains=search)
        )
    if role:
        queryset = queryset.filter(role__icontains=role)
    return queryset, role, search


def _unique_sections_for_school(school):
    if not school:
        return []
    return [
        item['section']
        for item in (
            StudentProfile.objects.filter(profile_category='school', college=school)
            .exclude(section='')
            .values('section')
            .distinct()
            .order_by('section')
        )
    ]


def _split_print_lines(value, width, max_lines):
    text = (value or '').strip()
    if not text:
        return [''] * max_lines

    lines = textwrap.wrap(
        text,
        width=width,
        max_lines=max_lines,
        placeholder='...',
        break_long_words=False,
        break_on_hyphens=False,
    )
    lines.extend([''] * (max_lines - len(lines)))
    return lines


def _format_member_dob(member):
    for attr in ('date_of_birth', 'dob', 'birth_date'):
        value = getattr(member, attr, None)
        if value:
            try:
                return value.strftime('%B %d, %Y')
            except AttributeError:
                return str(value)
    return ''


def _build_terms_lines(member_name, school_phone='', valid_till='', calendar_type='', custom_note=''):
    if custom_note and custom_note.strip():
        note = custom_note.strip()
    else:
        note = f"This card belongs to {member_name}. If found, please return it to the school administration office."
        if school_phone:
            note += f" Contact: {school_phone}."
        if valid_till:
            note += f" Valid till {valid_till}"
            if calendar_type:
                note += f" {calendar_type}"
            note += '.'
    return _split_print_lines(note, width=28, max_lines=5)


def _build_card_context(request, member, options=None):
    options = options or {}
    school = member.college
    selected_front_design = options.get('front_design') or _get_front_design_value(member)
    selected_back_design = options.get('back_design') or _get_back_design_value(member)
    selected_orientation = options.get('orientation') or (
        member.print_orientation if member.print_orientation in {'portrait', 'landscape'} else 'portrait'
    )
    selected_card_type = options.get('card_type') or member.print_card_type or 'id_card'
    selected_calendar = options.get('calendar') or (
        member.print_calendar if member.print_calendar in {'ad', 'bs'} else 'bs'
    )
    member_type_label = member.role or ('Teacher' if getattr(member, 'member_type', '') == 'teacher' else 'Student')
    member_address_lines = _split_print_lines(member.address, width=22, max_lines=2)
    member_dob = _format_member_dob(member)
    print_valid_till = options.get('valid_till') or member.print_valid_till

    return {
        'student': member,
        'school_name': school.name if school else '',
        'school_logo_url': school.logo.url if school and school.logo else '',
        'school_address': school.address if school else '',
        'school_phone': school.phone if school else '',
        'school_email': school.email if school else '',
        'school_slogan': school.slogan if school else '',
        'principal_name': school.principal_name if school else '',
        'principal_signature_url': school.principal_signature.url if school and school.principal_signature else '',
        'member_name': member.name,
        'member_type_label': member_type_label,
        'member_photo_url': member.profile_photo.url if member.profile_photo else '',
        'member_id_number': member.unique_identifier or member.roll_number or f'STU-{member.id}',
        'member_email': member.email,
        'member_phone': member.phone,
        'member_class_label': member.get_academic_level_display() if member.academic_level else '',
        'member_section': member.section,
        'member_roll_number': member.roll_number,
        'member_blood_group': member.blood_group,
        'member_emergency_contact_name': member.emergency_contact_name,
        'member_emergency_contact_phone': member.emergency_contact_phone,
        'member_map_url': member.map_url,
        'member_dob': member_dob,
        'member_address_lines': member_address_lines,
        'card_label': options.get('label') or member.print_label or PRINT_FRONT_THEMES[selected_front_design]['label'],
        'calendar_type': selected_calendar.upper(),
        'card_type_label': next((item['label'] for item in PRINT_CARD_TYPES if item['value'] == selected_card_type), 'ID Card'),
        'theme_primary': options.get('theme_primary') or (school.theme_primary if school and school.theme_primary else '#1a3a5c'),
        'theme_light_primary': options.get('theme_light_primary') or (school.theme_light_primary if school and school.theme_light_primary else '#f7f5f0'),
        'theme_secondary': options.get('theme_secondary') or (school.theme_secondary if school and school.theme_secondary else '#1a1a2a'),
        'theme_ternary': options.get('theme_ternary') or (school.theme_ternary if school and school.theme_ternary else '#c9a84c'),
        'display_organization': member.organization_name or (member.college.name if member.college else ''),
        'print_template': selected_front_design,
        'print_template_meta': PRINT_FRONT_THEMES[selected_front_design],
        'front_theme_key': selected_front_design,
        'front_theme': PRINT_FRONT_THEMES[selected_front_design],
        'back_theme_key': selected_back_design,
        'back_theme': PRINT_BACK_THEMES[selected_back_design],
        'print_orientation': selected_orientation,
        'print_card_type': selected_card_type,
        'print_calendar': selected_calendar,
        'print_valid_till': print_valid_till,
        'print_label': options.get('label') or member.print_label,
        'print_custom_note': member.print_custom_note,
        'terms_lines': _build_terms_lines(
            member.name,
            school.phone if school else '',
            print_valid_till,
            selected_calendar.upper(),
            member.print_custom_note,
        ),
        'contact_card_url': _build_public_contact_url(request, member),
        'nfc_url': _build_public_contact_url(request, member),
        'qr_code_url': reverse('print_qr_code', args=[member.id]),
    }


def _parse_print_options(request, school=None):
    front_design = request.POST.get('front_design') or request.GET.get('front_design') or 'front_official_wave'
    back_design = request.POST.get('back_design') or request.GET.get('back_design') or 'back_official_wave'
    if front_design not in STUDIO_FRONT_THEME_KEYS:
        front_design = 'front_official_wave'
    if back_design not in STUDIO_BACK_THEME_KEYS:
        back_design = 'back_official_wave'
    print_mode = request.POST.get('print_mode') or request.GET.get('print_mode') or ''
    if not print_mode:
        print_mode = 'a4' if (request.POST.get('a4_layout') or request.GET.get('a4_layout')) else 'single'
    if print_mode not in {'a4', 'single'}:
        print_mode = 'single'

    selected_front_theme = PRINT_FRONT_THEMES[front_design]
    selected_back_theme = PRINT_BACK_THEMES[back_design]
    default_palette = selected_front_theme.get('palette', [])
    theme_primary = request.POST.get('theme_primary') or request.GET.get('theme_primary') or (default_palette[0] if len(default_palette) > 0 else None) or (school.theme_primary if school else '#1a3a5c')
    theme_light_primary = request.POST.get('theme_light_primary') or request.GET.get('theme_light_primary') or (default_palette[1] if len(default_palette) > 1 else None) or (school.theme_light_primary if school else '#f7f5f0')
    theme_secondary = request.POST.get('theme_secondary') or request.GET.get('theme_secondary') or (default_palette[2] if len(default_palette) > 2 else None) or (school.theme_secondary if school else '#1a1a2a')
    theme_ternary = request.POST.get('theme_ternary') or request.GET.get('theme_ternary') or (default_palette[3] if len(default_palette) > 3 else None) or (school.theme_ternary if school else '#c9a84c')

    return {
        'front_design': front_design,
        'back_design': back_design,
        'orientation': request.POST.get('orientation') or request.GET.get('orientation') or 'portrait',
        'card_type': request.POST.get('card_type') or request.GET.get('card_type') or 'id_card',
        'calendar': request.POST.get('calendar') or request.GET.get('calendar') or 'bs',
        'valid_till': request.POST.get('valid_till') or request.GET.get('valid_till') or '',
        'label': request.POST.get('label') or request.GET.get('label') or '',
        'theme_primary': theme_primary,
        'theme_light_primary': theme_light_primary,
        'theme_secondary': theme_secondary,
        'theme_ternary': theme_ternary,
        'print_mode': print_mode,
        'a4_layout': print_mode == 'a4',
        'front_theme_label': selected_front_theme['label'],
        'back_theme_label': selected_back_theme['label'],
    }


def _filtered_print_queryset(request, school):
    queryset = _school_member_queryset(school)
    member_type = (request.POST.get('member_type') or request.GET.get('member_type') or 'all').strip()
    academic_level = (request.POST.get('academic_level') or request.GET.get('academic_level') or '').strip()
    section = (request.POST.get('section') or request.GET.get('section') or '').strip()
    search = (request.POST.get('q') or request.GET.get('q') or '').strip()

    if member_type in {'student', 'teacher'}:
        queryset = queryset.filter(member_type=member_type)
    if academic_level:
        queryset = queryset.filter(academic_level=academic_level)
    if section:
        queryset = queryset.filter(section__iexact=section)
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search)
            | Q(email__icontains=search)
            | Q(phone__icontains=search)
            | Q(role__icontains=search)
            | Q(roll_number__icontains=search)
        )
    return queryset, {
        'member_type': member_type,
        'academic_level': academic_level,
        'section': section,
        'q': search,
    }


def _resolve_selected_members(request, school):
    queryset, filter_state = _filtered_print_queryset(request, school)
    selected_ids = [value for value in request.POST.getlist('selected_ids') if value]
    if request.POST.get('select_filtered') == '1':
        members = list(queryset)
    else:
        members = list(queryset.filter(id__in=selected_ids))
    return members, filter_state


def _build_print_data_workbook(request, school, members):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Print Data'
    headers = [
        'QR Preview',
        'QR File',
        'Digital Profile URL',
        'Name',
        'DOB',
        'ID Number',
        'Member Type',
        'Class / Grade',
        'Section',
        'Roll Number',
        'Phone',
        'Email',
        'Blood Group',
        'Emergency Contact Name',
        'Emergency Contact Phone',
        'Address',
        'School Name',
        'School Phone',
        'School Address',
    ]
    sheet.append(headers)
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    qr_assets = []
    image_buffers = []
    for row_index, member in enumerate(members, start=2):
        profile_url = _build_public_contact_url(request, member)
        qr_bytes = _build_qr_png_bytes(profile_url)
        member_code = _safe_export_name(member.unique_identifier or member.roll_number or str(member.id), f'member_{member.id}')
        qr_filename = f'qr_codes/{member_code}_{_safe_export_name(member.name)}.png'
        qr_assets.append((qr_filename, qr_bytes))

        row = [
            '',
            qr_filename,
            profile_url,
            member.name,
            _format_member_dob(member),
            member.unique_identifier or member.roll_number or f'STU-{member.id}',
            member.get_member_type_display(),
            member.get_academic_level_display() if member.academic_level else '',
            member.section,
            member.roll_number,
            member.phone,
            member.email,
            member.blood_group,
            member.emergency_contact_name,
            member.emergency_contact_phone,
            member.address,
            school.name if school else '',
            school.phone if school else '',
            school.address if school else '',
        ]
        sheet.append(row)
        sheet.row_dimensions[row_index].height = 68
        qr_buffer = BytesIO(qr_bytes)
        image_buffers.append(qr_buffer)
        qr_image = OpenPyXLImage(qr_buffer)
        qr_image.width = 78
        qr_image.height = 78
        sheet.add_image(qr_image, f'A{row_index}')

    widths = [14, 34, 46, 26, 18, 22, 18, 18, 12, 16, 18, 28, 14, 26, 22, 34, 26, 18, 34]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    instructions = workbook.create_sheet('Printer Notes')
    instructions.append(['Purpose', 'This package gives the printing team raw data and QR assets. They can decide the physical card design.'])
    instructions.append(['QR behavior', 'Each QR opens the member digital profile / portfolio URL.'])
    instructions.append(['QR files', 'PNG files are included in the qr_codes folder inside this ZIP.'])
    instructions.append(['Suggested printed fields', 'Name, DOB, ID Number, Class / Grade, Section, QR code.'])
    instructions.append(['School', school.name if school else ''])
    instructions.append(['Exported at', timezone.now().strftime('%Y-%m-%d %H:%M')])
    instructions.column_dimensions['A'].width = 24
    instructions.column_dimensions['B'].width = 90
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        row[0].font = Font(bold=True)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), qr_assets


def _draw_card_front_pdf(pdf, card, x, y, width, height):
    primary = HexColor(card['theme_primary'])
    secondary = HexColor(card['theme_secondary'])
    tertiary = HexColor(card['theme_ternary'])
    light = HexColor(card['theme_light_primary'])

    pdf.setFillColor(white)
    pdf.roundRect(x, y, width, height, 14, fill=1, stroke=0)
    pdf.setFillColor(primary)
    pdf.roundRect(x, y + height - 45, width, 45, 14, fill=1, stroke=0)
    pdf.setFillColor(light)
    pdf.rect(x, y, width, 34, fill=1, stroke=0)

    pdf.setFillColor(white)
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawCentredString(x + width / 2, y + height - 27, card['school_name'][:34])
    pdf.setFont('Helvetica', 7)
    pdf.drawCentredString(x + width / 2, y + height - 38, (card['school_slogan'] or 'School Identity Platform')[:46])

    pdf.setFillColor(secondary)
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawCentredString(x + width / 2, y + height - 68, card['member_name'][:32])
    pdf.setFillColor(primary)
    pdf.setFont('Helvetica-Bold', 7)
    pdf.drawCentredString(x + width / 2, y + height - 80, card['member_type_label'][:28].upper())

    pdf.setFillColor(black)
    pdf.setFont('Helvetica-Bold', 8)
    details = [
        ('ID', card['member_id_number']),
        ('Email', card['member_email']),
        ('Phone', card['member_phone']),
    ]
    if card['member_class_label']:
        details.append(('Class', card['member_class_label']))
    if card['member_section']:
        details.append(('Section', card['member_section']))

    line_y = y + height - 100
    for label, value in details:
        if not value:
            continue
        pdf.drawString(x + 16, line_y, f'{label}:')
        pdf.setFont('Helvetica', 8)
        pdf.drawString(x + 54, line_y, str(value)[:28])
        pdf.setFont('Helvetica-Bold', 8)
        line_y -= 11

    pdf.setFillColor(tertiary)
    pdf.roundRect(x + 12, y + 10, width - 24, 12, 6, fill=1, stroke=0)


def _draw_card_back_pdf(pdf, card, x, y, width, height):
    primary = HexColor(card['theme_primary'])
    secondary = HexColor(card['theme_secondary'])
    tertiary = HexColor(card['theme_ternary'])
    light = HexColor(card['theme_light_primary'])

    pdf.setFillColor(white)
    pdf.roundRect(x, y, width, height, 14, fill=1, stroke=0)
    pdf.setFillColor(primary)
    pdf.roundRect(x, y + height - 45, width, 45, 14, fill=1, stroke=0)
    pdf.setFillColor(white)
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawCentredString(x + width / 2, y + height - 27, 'Scan For Profile')

    qr = qrcode.QRCode(version=1, box_size=5, border=1)
    qr.add_data(card['contact_card_url'])
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='black', back_color='white')
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    pdf.drawImage(ImageReader(qr_buffer), x + (width - 80) / 2, y + height - 160, 80, 80, preserveAspectRatio=True, mask='auto')

    pdf.setFillColor(secondary)
    pdf.setFont('Helvetica-Bold', 8)
    pdf.drawCentredString(x + width / 2, y + height - 172, 'School Contact')
    pdf.setFillColor(black)
    pdf.setFont('Helvetica', 8)
    if card['school_phone']:
        pdf.drawCentredString(x + width / 2, y + height - 184, str(card['school_phone'])[:32])
    if card['school_email']:
        pdf.drawCentredString(x + width / 2, y + height - 196, str(card['school_email'])[:32])

    pdf.setFillColor(light)
    pdf.roundRect(x + 12, y + 34, width - 24, 54, 8, fill=1, stroke=0)
    pdf.setFillColor(black)
    pdf.setFont('Helvetica-Bold', 8)
    pdf.drawString(x + 18, y + 76, 'Terms & Conditions')
    pdf.setFont('Helvetica', 7)
    note = f"This card belongs to {card['member_name']}. If found, please return it to the school office."
    pdf.drawString(x + 18, y + 64, note[:60])
    if card['print_valid_till']:
        pdf.drawString(x + 18, y + 52, f"Valid till {card['print_valid_till']} {card['calendar_type']}")

    pdf.setFillColor(tertiary)
    pdf.roundRect(x + 12, y + 10, width - 24, 12, 6, fill=1, stroke=0)


def _build_pdf_response(cards, print_mode, filename):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    pdf = canvas.Canvas(response, pagesize=A4)
    page_width, page_height = A4

    if print_mode == 'single':
        card_width = 220
        card_height = 350
        for card in cards:
            x = (page_width - card_width) / 2
            front_y = page_height - card_height - 40
            back_y = 40
            _draw_card_front_pdf(pdf, card, x, front_y, card_width, card_height)
            _draw_card_back_pdf(pdf, card, x, back_y, card_width, card_height)
            pdf.showPage()
    else:
        columns = 2
        card_width = 250
        card_height = 160
        horizontal_gap = 20
        vertical_gap = 20
        margin_x = (page_width - (columns * card_width) - horizontal_gap) / 2
        margin_top = 40
        rows_per_page = 4

        for side in ('front', 'back'):
            for index, card in enumerate(cards):
                page_index = index % (columns * rows_per_page)
                row = page_index // columns
                column = page_index % columns
                x = margin_x + column * (card_width + horizontal_gap)
                y = page_height - margin_top - ((row + 1) * card_height) - (row * vertical_gap)
                if side == 'front':
                    _draw_card_front_pdf(pdf, card, x, y, card_width, card_height)
                else:
                    _draw_card_back_pdf(pdf, card, x, y, card_width, card_height)
                if page_index == (columns * rows_per_page) - 1 or index == len(cards) - 1:
                    pdf.showPage()

    pdf.save()
    return response


def home(request):
    return render(request, 'home.html')


def _fallback_chat_reply(message):
    text = (message or '').lower()
    if 'price' in text or 'cost' in text or 'shipping' in text:
        return 'Our NFC card pricing and delivery options depend on the card type, quantity, and delivery location. You can check the pricing section or share your quantity and location here so our team can follow up.'
    if 'nfc' in text or 'phone' in text or 'compatible' in text:
        return 'SkillConnect cards work by opening a digital contact card when tapped on an NFC-enabled phone. If a phone does not support NFC, the printed QR code can still open the same digital profile.'
    if 'school' in text or 'student' in text or 'teacher' in text or 'id card' in text:
        return 'For schools, we can create student and teacher digital profiles, generate QR/NFC-linked ID cards, and print either A4 batches or individual card layouts.'
    if 'update' in text or 'change' in text or 'edit' in text:
        return 'Yes, profile information can be updated later from the dashboard. The NFC tap and QR code continue pointing to the live digital profile, so the printed card does not need to be changed for basic profile updates.'
    return 'Thanks for the question. SkillConnect helps people and schools share contact details through NFC cards, QR codes, and live digital profiles. Please share what you want to do, and I can guide you.'


@require_POST
def ai_chat(request):
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid chat request.'}, status=400)

    email = (payload.get('email') or '').strip()
    message = (payload.get('message') or '').strip()
    if not email or not message:
        return JsonResponse({'error': 'Email and message are required.'}, status=400)
    if len(message) > 1200:
        return JsonResponse({'error': 'Please keep your question under 1200 characters.'}, status=400)

    fallback_reply = _fallback_chat_reply(message)
    if not settings.OPENAI_API_KEY:
        return JsonResponse({
            'reply': fallback_reply,
            'ai_enabled': False,
        })

    try:
        response = requests.post(
            'https://api.openai.com/v1/responses',
            headers={
                'Authorization': f'Bearer {settings.OPENAI_API_KEY}',
                'Content-Type': 'application/json',
            },
            json={
                'model': settings.OPENAI_CHAT_MODEL,
                'input': [
                    {
                        'role': 'system',
                        'content': (
                            'You are the SkillConnect website assistant. Answer customers clearly and briefly. '
                            'SkillConnect sells NFC-enabled cards, QR-linked digital contact cards, school ID card '
                            'printing tools, student/teacher profile dashboards, and live profile updates. '
                            'Do not promise exact pricing or delivery dates unless the customer provides details. '
                            'Ask for contact details or tell them the team can follow up when needed.'
                        ),
                    },
                    {
                        'role': 'user',
                        'content': f'Customer email: {email}\nCustomer question: {message}',
                    },
                ],
                'temperature': 0.4,
                'max_output_tokens': 220,
            },
            timeout=20,
        )
        response.raise_for_status()
        data = response.json()
        reply = (data.get('output_text') or '').strip()
        if not reply:
            output = data.get('output') or []
            reply = ''.join(
                part.get('text', '')
                for item in output
                for part in item.get('content', [])
                if part.get('type') in {'output_text', 'text'}
            ).strip()
        return JsonResponse({'reply': reply or fallback_reply, 'ai_enabled': True})
    except requests.RequestException:
        return JsonResponse({'reply': fallback_reply, 'ai_enabled': False})


def dashboard_login(request):
    if request.user.is_authenticated:
        return redirect('admin_dashboard')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is None:
            messages.error(request, 'Invalid username or password.')
        else:
            role = _get_user_role(user)
            if role == 'public':
                messages.error(request, 'This account does not have dashboard access.')
                return render(request, 'auth/login.html')
            login(request, user)
            return redirect('admin_dashboard')
    return render(request, 'auth/login.html')


@login_required
def dashboard_logout(request):
    logout(request)
    messages.success(request, 'You have been logged out.')
    return redirect('dashboard_login')


def _build_tracked_action_url(student_id, action):
    return f'/student/{student_id}/action/{action}/'


def _log_profile_activity(student, event_type, action=''):
    ProfileActivity.objects.create(student=student, event_type=event_type, action=action)


def _normalize_public_url(url):
    if url and not url.startswith(('http://', 'https://')):
        return f'https://{url}'
    return url or ''


def _build_public_contact_url(request, student):
    return request.build_absolute_uri(reverse('student_contact_card', args=[student.id]))


def _safe_export_name(value, fallback='member'):
    cleaned = re.sub(r'[^A-Za-z0-9_-]+', '_', value or '').strip('_')
    return cleaned[:80] or fallback


def _build_qr_png_bytes(url):
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    image = qr.make_image(fill_color='black', back_color='white')
    buffer = BytesIO()
    image.save(buffer, format='PNG')
    return buffer.getvalue()


def _media_url(file_field):
    return file_field.url if file_field else ''


def _google_maps_url(address):
    if not address:
        return ''
    return f'https://www.google.com/maps/search/?api=1&{urlencode({"query": address})}'


def _school_card_context(request, student):
    school = student.college
    school_name = school.name if school else ''
    school_address = school.address if school and school.address else ''
    school_website_url = _normalize_public_url(school.website) if school else ''
    school_phone = school.phone if school else ''
    school_logo_url = _media_url(school.logo) if school else ''
    school_map_url = _google_maps_url(school_address)

    member_type_label = student.get_member_type_display() if student.member_type else 'Student'
    grade_label = student.get_academic_level_display() if student.academic_level else (student.role or member_type_label)
    grade_section = grade_label
    if student.section:
        grade_section = f'{grade_label} - Section {student.section}'

    website_url = _normalize_public_url(student.website) or school_website_url
    direct_map_url = _normalize_public_url(student.map_url) or _google_maps_url(student.address) or school_map_url
    navigate_url = _build_tracked_action_url(student.id, 'map') if (student.map_url or student.address) else direct_map_url

    social_links = []
    if student.instagram:
        social_links.append({'key': 'instagram', 'url': _build_tracked_action_url(student.id, 'social-instagram'), 'label': 'Instagram', 'icon': 'instagram'})
    if student.facebook:
        social_links.append({'key': 'facebook', 'url': _build_tracked_action_url(student.id, 'social-facebook'), 'label': 'Facebook', 'icon': 'facebook'})
    if student.whatsapp:
        social_links.append({'key': 'whatsapp', 'url': _build_tracked_action_url(student.id, 'whatsapp'), 'label': 'WhatsApp', 'icon': 'message-circle'})
    if student.linkedin:
        social_links.append({'key': 'linkedin', 'url': _build_tracked_action_url(student.id, 'social-linkedin'), 'label': 'LinkedIn', 'icon': 'linkedin'})

    return {
        'student': student,
        'school_name': school_name,
        'school_website': school.website if school and school.website else '',
        'school_website_url': school_website_url,
        'school_phone': school_phone,
        'school_address': school_address,
        'school_logo_url': school_logo_url,
        'cover_photo_url': _media_url(student.cover_photo),
        'student_photo_url': _media_url(student.profile_photo),
        'member_type_label': member_type_label,
        'grade_label': grade_label,
        'section_label': student.section,
        'grade_section': grade_section,
        'student_identifier': student.unique_identifier or student.roll_number or f'STU-{student.id}',
        'parent_label': 'Parent / Guardian' if student.member_type == 'student' else 'Emergency Contact',
        'parent_name': student.emergency_contact_name,
        'emergency_contact_name': student.emergency_contact_name,
        'emergency_contact_phone': student.emergency_contact_phone,
        'blood_group': student.blood_group,
        'student_address': student.address,
        'website_url': website_url,
        'map_url': direct_map_url,
        'navigate_url': navigate_url,
        'phone_action_url': _build_tracked_action_url(student.id, 'phone') if student.phone else '',
        'whatsapp_action_url': _build_tracked_action_url(student.id, 'whatsapp') if student.whatsapp else '',
        'download_vcard_url': reverse('download_vcard', args=[student.id]),
        'qr_code_url': reverse('print_qr_code', args=[student.id]),
        'public_card_url': _build_public_contact_url(request, student),
        'social_links': social_links,
    }


def _extract_role_from_post(request, profile_category, fallback=''):
    if profile_category == 'school':
        return request.POST.get('school_role') or request.POST.get('role') or fallback
    return request.POST.get('role') or fallback


def _assign_profile_skills(student, request):
    selected_ids = request.POST.getlist('skills')
    selected_skills = list(Skill.objects.filter(id__in=selected_ids))

    raw_custom_skills = request.POST.get('custom_skills', '')
    custom_names = []
    for part in raw_custom_skills.replace('\n', ',').split(','):
        cleaned = part.strip()
        if cleaned:
            custom_names.append(cleaned)

    for name in custom_names:
        skill, _ = Skill.objects.get_or_create(name=name)
        selected_skills.append(skill)

    unique_skills = []
    seen_ids = set()
    for skill in selected_skills:
        if skill.id not in seen_ids:
            unique_skills.append(skill)
            seen_ids.add(skill.id)

    student.skills.set(unique_skills)


def _get_front_design_value(student):
    candidate = getattr(student, 'print_front_design', '') or student.print_template
    candidate = LEGACY_PRINT_TEMPLATE_MAP.get(candidate, candidate)
    if candidate not in PRINT_FRONT_THEMES:
        return 'front_academic_blue'
    return candidate


def _get_back_design_value(student):
    candidate = getattr(student, 'print_back_design', '')
    if candidate not in PRINT_BACK_THEMES:
        return 'back_qr_clean'
    return candidate


def _get_print_template_label(student):
    front_key = _get_front_design_value(student)
    back_key = _get_back_design_value(student)
    return f"{PRINT_FRONT_THEMES[front_key]['label']} / {PRINT_BACK_THEMES[back_key]['label']}"


def _is_faculty_role(role_name):
    normalized = (role_name or '').strip().lower()
    return any(keyword in normalized for keyword in FACULTY_ROLE_KEYWORDS)


def _extract_member_type_from_post(request, fallback='student'):
    value = request.POST.get('member_type') or fallback
    if value in {'student', 'teacher', 'other'}:
        return value
    return fallback


def _calculate_profile_completion(student):
    checks = [
        bool(student.name),
        bool(student.username),
        bool(student.email),
        bool(student.phone),
        bool(student.role),
        bool(student.bio),
        bool(student.address),
        bool(student.profile_photo),
        bool(student.website),
        bool(student.social_stack),
    ]
    completed = sum(1 for item in checks if item)
    return int((completed / len(checks)) * 100)


def _is_school_faculty_profile(student):
    if getattr(student, 'member_type', '') == 'teacher':
        return True
    role = (student.role or '').strip().lower()
    if student.profile_category != 'school':
        return False
    return any(keyword in role for keyword in FACULTY_ROLE_KEYWORDS)


def _get_portfolio_template(student, requested_template=None):
    allowed_templates = PORTFOLIO_TEMPLATES.get(student.user_type, [])
    if not allowed_templates:
        return None

    if requested_template in allowed_templates:
        return requested_template

    if _is_school_faculty_profile(student) and 'profile6.html' in allowed_templates:
        return 'profile6.html'

    current_template = student.portfolio_template
    if current_template in allowed_templates:
        return current_template

    return allowed_templates[0]


def profile(request, student_id):
    return redirect('contact_card', student_id=student_id)


def print_qr_code(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    return HttpResponse(_build_qr_png_bytes(_build_public_contact_url(request, student)), content_type='image/png')


@login_required
def print_card_preview(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    permission_response = _require_profile_access(request, student)
    if permission_response:
        return permission_response
    selected_front_design = _get_front_design_value(student)
    selected_back_design = _get_back_design_value(student)
    selected_orientation = student.print_orientation if student.print_orientation in {'portrait', 'landscape'} else 'portrait'
    selected_card_type = student.print_card_type or 'id_card'
    selected_calendar = student.print_calendar if student.print_calendar in {'ad', 'bs'} else 'bs'

    school = student.college
    member_type_label = student.role or ('Teacher' if getattr(student, 'member_type', '') == 'teacher' else 'Student')
    member_address_lines = _split_print_lines(student.address, width=22, max_lines=2)
    member_dob = _format_member_dob(student)
    context = {
        'student': student,
        'school_name': school.name if school else '',
        'school_logo_url': school.logo.url if school and school.logo else '',
        'school_address': school.address if school else '',
        'school_phone': school.phone if school else '',
        'school_email': school.email if school else '',
        'school_slogan': school.slogan if school else '',
        'principal_name': school.principal_name if school else '',
        'principal_signature_url': school.principal_signature.url if school and school.principal_signature else '',
        'member_name': student.name,
        'member_type_label': member_type_label,
        'member_photo_url': student.profile_photo.url if student.profile_photo else '',
        'member_id_number': student.unique_identifier or student.roll_number or f'STU-{student.id}',
        'member_email': student.email,
        'member_phone': student.phone,
        'member_class_label': student.get_academic_level_display() if student.academic_level else '',
        'member_section': student.section,
        'member_roll_number': student.roll_number,
        'member_blood_group': student.blood_group,
        'member_emergency_contact_name': student.emergency_contact_name,
        'member_emergency_contact_phone': student.emergency_contact_phone,
        'member_map_url': student.map_url,
        'member_dob': member_dob,
        'member_address_lines': member_address_lines,
        'card_label': student.print_label or PRINT_FRONT_THEMES[selected_front_design]['label'],
        'calendar_type': selected_calendar.upper(),
        'card_type_label': next((item['label'] for item in PRINT_CARD_TYPES if item['value'] == selected_card_type), 'ID Card'),
        'theme_primary': school.theme_primary if school and school.theme_primary else '#1a3a5c',
        'theme_light_primary': school.theme_light_primary if school and school.theme_light_primary else '#f7f5f0',
        'theme_secondary': school.theme_secondary if school and school.theme_secondary else '#1a1a2a',
        'theme_ternary': school.theme_ternary if school and school.theme_ternary else '#c9a84c',
        'display_organization': student.organization_name or (student.college.name if student.college else ''),
        'print_template': selected_front_design,
        'print_template_meta': PRINT_FRONT_THEMES[selected_front_design],
        'front_theme_key': selected_front_design,
        'front_theme': PRINT_FRONT_THEMES[selected_front_design],
        'back_theme_key': selected_back_design,
        'back_theme': PRINT_BACK_THEMES[selected_back_design],
        'print_orientation': selected_orientation,
        'print_card_type': selected_card_type,
        'print_calendar': selected_calendar,
        'print_valid_till': student.print_valid_till,
        'print_label': student.print_label,
        'print_custom_note': student.print_custom_note,
        'terms_lines': _build_terms_lines(
            student.name,
            school.phone if school else '',
            student.print_valid_till,
            selected_calendar.upper(),
            student.print_custom_note,
        ),
        'contact_card_url': _build_public_contact_url(request, student),
        'nfc_url': _build_public_contact_url(request, student),
        'qr_code_url': reverse('print_qr_code', args=[student.id]),
    }
    return render(request, 'print/preview.html', context)


@login_required
def admin_dashboard(request):
    role = _get_user_role(request.user)
    if role == 'teacher':
        owned_profile = _get_owned_profile(request.user)
        if owned_profile:
            request.session[_student_edit_session_key(owned_profile.id)] = True
            return redirect('student_owner_dashboard', student_id=owned_profile.id)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to the admin dashboard.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=False)
    school_members = _school_member_queryset(school) if school else StudentProfile.objects.none()

    if role == 'super_admin':
        total_schools = College.objects.count()
        total_students = StudentProfile.objects.filter(profile_category='school', member_type='student').count()
        total_teachers = StudentProfile.objects.filter(profile_category='school', member_type='teacher').count()
        recent_activities = ProfileActivity.objects.select_related('student', 'student__college')[:8]
    else:
        total_schools = 1 if school else 0
        total_students = school_members.filter(member_type='student').count() if school else 0
        total_teachers = school_members.filter(member_type='teacher').count() if school else 0
        recent_activities = ProfileActivity.objects.select_related('student', 'student__college').filter(
            student__college=school
        )[:8] if school else ProfileActivity.objects.none()

    dashboard_features = [
        {
            'title': 'Student Records',
            'description': 'Add, filter, edit, reset passwords, and open public contact cards.',
            'icon': 'graduation-cap',
            'url': f"{reverse('dashboard_students')}{_build_dashboard_query(school)}" if school else '',
        },
        {
            'title': 'Teachers / Staff',
            'description': 'Manage staff profiles, dashboard access, and school-facing contact details.',
            'icon': 'briefcase',
            'url': f"{reverse('dashboard_teachers')}{_build_dashboard_query(school)}" if school else '',
        },
        {
            'title': 'ID Card Studio',
            'description': 'Batch select members, preview cards, and export print-ready PDFs.',
            'icon': 'printer',
            'url': f"{reverse('dashboard_print')}{_build_dashboard_query(school)}" if school else '',
        },
        {
            'title': 'QR Data Export',
            'description': 'Filter members and download Excel data with QR PNG assets for outside printing.',
            'icon': 'qr-code',
            'url': f"{reverse('dashboard_qr_export')}{_build_dashboard_query(school)}" if school else '',
        },
        {
            'title': 'Bulk Upload',
            'description': 'Import students or staff from CSV/XLSX files with school defaults applied.',
            'icon': 'upload-cloud',
            'url': f"{reverse('dashboard_bulk_upload')}{_build_dashboard_query(school)}" if school else '',
        },
        {
            'title': 'School Settings',
            'description': 'Update branding, admin login, principal details, and print colors.',
            'icon': 'settings',
            'url': f"{reverse('dashboard_settings')}{_build_dashboard_query(school)}" if school else '',
        },
    ]
    if role == 'super_admin':
        dashboard_features.insert(0, {
            'title': 'Super Admin Controls',
            'description': 'Create schools, assign school admins, switch workspaces, and remove schools.',
            'icon': 'shield-check',
            'url': reverse('admin_dashboard'),
        })

    context = {
        **_school_dashboard_context(request, 'home', school, schools),
        'total_schools': total_schools,
        'total_students': total_students,
        'total_teachers': total_teachers,
        'school_student_count': school_members.filter(member_type='student').count() if school else 0,
        'school_teacher_count': school_members.filter(member_type='teacher').count() if school else 0,
        'school_member_count': school_members.count() if school else 0,
        'recent_activities': recent_activities,
        'schools': schools,
        'dashboard_features': dashboard_features,
    }
    return render(request, 'dashboard/home.html', context)


@login_required
def dashboard_students(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to the students module.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Create a school first to manage students.')
        return redirect('admin_dashboard')

    students = _school_member_queryset(school, 'student')
    students, academic_level, section, search = _apply_student_filters(students, request)
    context = {
        **_school_dashboard_context(request, 'students', school, schools),
        'students': students,
        'academic_level': academic_level,
        'section_filter': section,
        'search_query': search,
        'sections': _unique_sections_for_school(school),
        'academic_level_choices': [{'value': value, 'label': label} for value, label in ACADEMIC_LEVEL_CHOICES],
    }
    return render(request, 'dashboard/students.html', context)


@login_required
def dashboard_teachers(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to the teachers module.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Create a school first to manage teachers.')
        return redirect('admin_dashboard')

    teacher_base_queryset = _school_member_queryset(school, 'teacher')
    teachers, role_filter, search = _apply_teacher_filters(teacher_base_queryset, request)
    role_options = [
        role for role in teacher_base_queryset.exclude(role='').values_list('role', flat=True).distinct().order_by('role')
    ]
    admin_role_keywords = ('principal', 'vice principal', 'coordinator', 'head', 'dean', 'chairman', 'chairperson')
    teacher_role_keywords = ('teacher', 'lecturer', 'professor', 'mentor', 'faculty', 'instructor', 'trainer')
    admin_count = 0
    teacher_count = 0
    other_count = 0
    for member in teacher_base_queryset:
        role_name = (member.role or '').strip().lower()
        if any(keyword in role_name for keyword in admin_role_keywords):
            admin_count += 1
        elif any(keyword in role_name for keyword in teacher_role_keywords) or not role_name:
            teacher_count += 1
        else:
            other_count += 1
    context = {
        **_school_dashboard_context(request, 'teachers', school, schools),
        'teachers': teachers,
        'role_filter': role_filter,
        'search_query': search,
        'role_options': role_options,
        'total_teacher_records': teacher_base_queryset.count(),
        'teacher_count': teacher_count,
        'admin_count': admin_count,
        'other_count': other_count,
    }
    return render(request, 'dashboard/teachers.html', context)


@login_required
def dashboard_settings(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to school settings.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Create a school first to open school settings.')
        return redirect('admin_dashboard')

    if request.method == 'POST':
        school.name = request.POST.get('name', school.name)
        school.slogan = request.POST.get('slogan', school.slogan)
        school.address = request.POST.get('address', school.address)
        school.principal_name = request.POST.get('principal_name', school.principal_name)
        school.website = request.POST.get('website', school.website)
        school.email = request.POST.get('email', school.email)
        school.phone = request.POST.get('phone', school.phone)
        school.theme_primary = request.POST.get('theme_primary', school.theme_primary or '#1a3a5c')
        school.theme_light_primary = request.POST.get('theme_light_primary', school.theme_light_primary or '#f7f5f0')
        school.theme_secondary = request.POST.get('theme_secondary', school.theme_secondary or '#1a1a2a')
        school.theme_ternary = request.POST.get('theme_ternary', school.theme_ternary or '#c9a84c')
        school.description = request.POST.get('description', school.description)
        if request.FILES.get('logo'):
            school.logo = request.FILES['logo']
        if request.FILES.get('principal_signature'):
            school.principal_signature = request.FILES['principal_signature']
        admin_username = request.POST.get('admin_username', '').strip()
        admin_password = request.POST.get('admin_password', '').strip()
        if admin_username:
            _sync_school_admin_user(school, admin_username, admin_password)
        school.save()
        messages.success(request, 'School settings updated successfully.')
        return redirect(f"{reverse('dashboard_settings')}{_build_dashboard_query(school)}")

    context = {
        **_school_dashboard_context(request, 'settings', school, schools),
        'college': school,
        'admin_username': school.admin_user.username if school.admin_user else '',
    }
    return render(request, 'dashboard/settings.html', context)


@login_required
def dashboard_bulk_upload(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to bulk upload.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=False)
    upload_summary = None

    if request.method == 'POST' and request.FILES.get('file'):
        try:
            import pandas as pd
        except ImportError:
            messages.error(request, 'Bulk upload requires pandas.')
            return redirect('dashboard_bulk_upload')

        upload_file = request.FILES['file']
        role_type = request.POST.get('role_type', 'student')
        selected_school = school
        if _is_super_admin(request.user):
            school_id = (request.POST.get('school') or '').strip()
            selected_school = College.objects.filter(id=school_id).first() if school_id else school

        if not selected_school:
            messages.error(request, 'Select a school before running bulk upload.')
            return redirect('dashboard_bulk_upload')

        filename = upload_file.name.lower()
        try:
            if filename.endswith('.csv'):
                dataframe = pd.read_csv(upload_file)
            else:
                dataframe = pd.read_excel(upload_file)
        except Exception as exc:
            messages.error(request, f'Could not read the uploaded file: {exc}')
            return redirect(f"{reverse('dashboard_bulk_upload')}{_build_dashboard_query(selected_school)}")

        required_columns = ['name', 'phone']
        missing_columns = [col for col in required_columns if col not in dataframe.columns]
        if missing_columns:
            messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
            return redirect(f"{reverse('dashboard_bulk_upload')}{_build_dashboard_query(selected_school)}")

        created_count = 0
        skipped_rows = []
        for index, row in dataframe.fillna('').iterrows():
            name = str(row.get('name', '')).strip()
            phone = str(row.get('phone', '')).strip()
            if not name or not phone:
                skipped_rows.append(index + 2)
                continue

            username_seed = str(row.get('username') or name.lower().replace(' ', ''))[:120]
            raw_password = _generate_profile_password(name)
            student = StudentProfile(
                name=name,
                phone=phone,
                email=str(row.get('email', '')).strip(),
                username=_ensure_unique_username(username_seed or f'member{index + 1}'),
                college=selected_school,
                profile_category='school',
                member_type=role_type,
                role=str(row.get('role', 'Teacher' if role_type == 'teacher' else 'Student')).strip(),
                address=str(row.get('address', '')).strip(),
                emergency_contact_name=str(row.get('emergency_contact_name', '')).strip(),
                emergency_contact_phone=str(row.get('emergency_contact_phone', '')).strip(),
                map_url=str(row.get('map_url', '')).strip() or None,
                academic_level=str(row.get('academic_level', '')).strip(),
                section=str(row.get('section', '')).strip(),
                roll_number=str(row.get('roll_number', '')).strip(),
                blood_group=str(row.get('blood_group', '')).strip(),
                gender=str(row.get('gender', '')).strip(),
                organization_name=selected_school.name,
                password=raw_password,
            )
            student.save()
            if _profile_supports_self_service(student):
                _sync_profile_auth_user(student, raw_password)
                student.save(update_fields=['auth_user', 'username'])
            created_count += 1

        upload_summary = {
            'created_count': created_count,
            'skipped_rows': skipped_rows,
            'role_type': role_type,
            'school': selected_school,
            'filename': upload_file.name,
        }
        messages.success(request, f'Bulk upload completed. Created {created_count} {role_type} records.')
        school = selected_school

    context = {
        **_school_dashboard_context(request, 'bulk_upload', school, schools),
        'college': school,
        'upload_summary': upload_summary,
    }
    return render(request, 'dashboard/bulk_upload.html', context)


@login_required
def dashboard_print(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to the ID card module.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Create a school first to use the print module.')
        return redirect('admin_dashboard')

    members, filter_state = _filtered_print_queryset(request, school)
    print_options = _parse_print_options(request, school)
    preview_member = members.first()
    preview_card = _build_card_context(request, preview_member, print_options) if preview_member else None
    preview_label = (
        print_options.get('label')
        or (preview_card.get('card_label') if preview_card else '')
        or 'Student Identity Card'
    )
    context = {
        **_school_dashboard_context(request, 'print', school, schools),
        'college': school,
        'members': members,
        'member_count': members.count(),
        'school_student_total': _school_member_queryset(school, 'student').count(),
        'school_teacher_total': _school_member_queryset(school, 'teacher').count(),
        'filter_state': filter_state,
        'sections': _unique_sections_for_school(school),
        'academic_level_choices': [{'value': value, 'label': label} for value, label in ACADEMIC_LEVEL_CHOICES],
        'print_front_design_choices': [
            {
                'value': key,
                'label': PRINT_FRONT_THEMES[key]['label'],
                'description': PRINT_FRONT_THEMES[key]['description'],
                'palette': PRINT_FRONT_THEMES[key]['palette'],
            }
            for key in STUDIO_FRONT_THEME_KEYS
        ],
        'print_back_design_choices': [
            {
                'value': key,
                'label': PRINT_BACK_THEMES[key]['label'],
                'description': PRINT_BACK_THEMES[key]['description'],
                'palette': PRINT_BACK_THEMES[key]['palette'],
            }
            for key in STUDIO_BACK_THEME_KEYS
        ],
        'print_orientations': PRINT_ORIENTATIONS,
        'print_card_types': PRINT_CARD_TYPES,
        'print_options': print_options,
        'preview_card': preview_card,
        'preview_label': preview_label,
    }
    return render(request, 'dashboard/print.html', context)


@login_required
def dashboard_qr_export(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to QR data export.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Create a school first to export QR data.')
        return redirect('admin_dashboard')

    members, filter_state = _filtered_print_queryset(request, school)
    context = {
        **_school_dashboard_context(request, 'qr_export', school, schools),
        'college': school,
        'members': members,
        'member_count': members.count(),
        'school_student_total': _school_member_queryset(school, 'student').count(),
        'school_teacher_total': _school_member_queryset(school, 'teacher').count(),
        'filter_state': filter_state,
        'sections': _unique_sections_for_school(school),
        'academic_level_choices': [{'value': value, 'label': label} for value, label in ACADEMIC_LEVEL_CHOICES],
    }
    return render(request, 'dashboard/qr_export.html', context)


@login_required
def dashboard_print_preview(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to print preview.')
        return redirect('dashboard_login')

    school, schools = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Select a school before previewing cards.')
        return redirect('admin_dashboard')

    members, filter_state = _resolve_selected_members(request, school)
    if not members:
        messages.error(request, 'Select at least one student or teacher before previewing.')
        return redirect(f"{reverse('dashboard_print')}{_build_dashboard_query(school)}")

    options = _parse_print_options(request, school)
    cards = [_build_card_context(request, member, options) for member in members]
    template_name = 'print/a4_print.html' if options['print_mode'] == 'a4' else 'print/single_card.html'
    context = {
        **_school_dashboard_context(request, 'print', school, schools),
        'college': school,
        'cards': cards,
        'selected_members': members,
        'selected_count': len(cards),
        'filter_state': filter_state,
        'print_options': options,
        'selected_ids': [member.id for member in members],
    }
    return render(request, template_name, context)


@login_required
def dashboard_print_export_pdf(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to PDF export.')
        return redirect('dashboard_login')

    school, _ = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Select a school before exporting cards.')
        return redirect('admin_dashboard')

    members, _ = _resolve_selected_members(request, school)
    if not members:
        messages.error(request, 'Select at least one student or teacher before exporting.')
        return redirect(f"{reverse('dashboard_print')}{_build_dashboard_query(school)}")

    options = _parse_print_options(request, school)
    cards = [_build_card_context(request, member, options) for member in members]
    filename = f"{school.name.replace(' ', '_').lower()}_id_cards.pdf"
    return _build_pdf_response(cards, options['print_mode'], filename)


@login_required
def dashboard_qr_export_download(request):
    role = _get_user_role(request.user)
    if role not in {'super_admin', 'school_admin'}:
        messages.error(request, 'You do not have access to QR data export.')
        return redirect('dashboard_login')

    school, _ = _resolve_dashboard_school(request, required=True)
    if not school:
        messages.error(request, 'Select a school before exporting QR data.')
        return redirect('admin_dashboard')

    members, _ = _resolve_selected_members(request, school)
    if not members:
        messages.error(request, 'Select at least one student or teacher before exporting QR data.')
        return redirect(f"{reverse('dashboard_qr_export')}{_build_dashboard_query(school)}")

    workbook_bytes, qr_assets = _build_print_data_workbook(request, school, members)
    school_slug = _safe_export_name(school.name, 'school').lower()
    filename = f'{school_slug}_qr_print_data.zip'
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w', ZIP_DEFLATED) as archive:
        archive.writestr(f'{school_slug}_print_data.xlsx', workbook_bytes)
        archive.writestr('README.txt', (
            'QR print data package\n\n'
            'Use the Excel file for member data. Use the PNG files in qr_codes/ for card artwork.\n'
            'Each QR opens the public digital profile for that student, teacher, or staff member.\n'
        ))
        for qr_filename, qr_bytes in qr_assets:
            archive.writestr(qr_filename, qr_bytes)

    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def add_user(request):
    messages.error(request, 'Independent profile creation is disabled in the school identity platform.')
    return redirect('admin_dashboard')


def edit_student_auth(request, student_id):
    student = get_object_or_404(StudentProfile, id=student_id)
    if not _profile_supports_self_service(student):
        messages.error(request, 'Students are managed by the school. Please contact your school administrator for profile changes.')
        return redirect('contact_card', student_id=student.id)
    if _can_manage_profile(request.user, student):
        request.session[_student_edit_session_key(student.id)] = True
        return redirect('student_owner_dashboard', student_id=student.id)
    if _is_student_edit_authorized(request, student.id):
        return redirect('student_owner_dashboard', student_id=student.id)
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        if username == student.username and check_password(password, student.password):
            request.session[_student_edit_session_key(student.id)] = True
            messages.success(request, 'Login successful. You can now manage this profile.')
            return redirect('student_owner_dashboard', student_id=student.id)
        messages.error(request, 'Invalid username or password.')
    return render(request, 'edit_student_auth.html', {'student': student})


def logout_student_edit(request, student_id):
    request.session.pop(_student_edit_session_key(student_id), None)
    messages.success(request, 'You have been logged out from teacher profile management.')
    return redirect('contact_card', student_id=student_id)


@login_required
def edit_student(request, student_id):
    student = get_object_or_404(StudentProfile, id=student_id)
    permission_response = _require_profile_access(request, student)
    if permission_response:
        return permission_response
    request.session[_student_edit_session_key(student_id)] = True
    messages.success(request, 'Admin access enabled for this profile.')
    return redirect('edit_student_manual', student_id=student_id)


@login_required
def reset_student_password(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    permission_response = _require_profile_access(request, student)
    if permission_response:
        return permission_response
    raw_password = _generate_profile_password(student.name)
    student.password = raw_password
    student.save(update_fields=['password'])
    _sync_profile_auth_user(student, raw_password)
    student.save(update_fields=['password', 'auth_user', 'username'])
    if _profile_supports_self_service(student):
        messages.success(request, f'{student.name} password reset. New password: {raw_password}')
    else:
        messages.success(request, f'{student.name} password reset. This profile remains school-managed.')
    redirect_target = request.POST.get('next') or reverse('admin_dashboard')
    return redirect(redirect_target)


@login_required
def delete_student_profile(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    permission_response = _require_profile_access(request, student)
    if permission_response:
        return permission_response
    name = student.name
    student.delete()
    messages.success(request, f'{name} deleted successfully.')
    return redirect(request.POST.get('next') or reverse('admin_dashboard'))


@login_required
def student_owner_dashboard(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    if not _profile_supports_self_service(student):
        messages.error(request, 'Students do not have self-service dashboard access.')
        return redirect('contact_card', student_id=student.id)
    if not (_can_manage_profile(request.user, student) or _is_student_edit_authorized(request, student.id)):
        messages.error(request, 'Please log in with your profile username and password to manage this profile.')
        return redirect('edit_student_auth', student_id=student.id)

    if request.method == 'POST':
        action = request.POST.get('dashboard_action')
        if action == 'toggle_contact_card':
            student.show_contact_card = not student.show_contact_card
            student.save(update_fields=['show_contact_card'])
            messages.success(request, f"Contact card visibility {'enabled' if student.show_contact_card else 'disabled'}.")
            return redirect('student_owner_dashboard', student_id=student.id)
        if action == 'change_password':
            current_password = request.POST.get('current_password', '')
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')

            if not check_password(current_password, student.password):
                messages.error(request, 'Current password is incorrect.')
            elif len(new_password) < 6:
                messages.error(request, 'New password must be at least 6 characters long.')
            elif new_password != confirm_password:
                messages.error(request, 'New password and confirm password do not match.')
            else:
                student.password = new_password
                student.save(update_fields=['password'])
                _sync_profile_auth_user(student, new_password)
                student.save(update_fields=['password', 'auth_user', 'username'])
                messages.success(request, 'Password updated successfully.')
            return redirect('student_owner_dashboard', student_id=student.id)

    display_organization = student.organization_name or (student.college.name if student.college else 'Personal Brand')
    total_engagement = student.views + student.downloads + student.contact_clicks
    profile_completion = _calculate_profile_completion(student)
    seven_days_ago = timezone.now() - timedelta(days=6)
    activity_rows = (
        student.activities.filter(created_at__date__gte=seven_days_ago.date())
        .annotate(day=TruncDate('created_at'))
        .values('day', 'event_type')
        .annotate(total=Count('id'))
        .order_by('day', 'event_type')
    )
    activity_map = {}
    for row in activity_rows:
        activity_map.setdefault(row['day'], {'view': 0, 'download': 0, 'contact': 0})
        activity_map[row['day']][row['event_type']] = row['total']

    daily_analytics = []
    for offset in range(6, -1, -1):
        day = (timezone.now() - timedelta(days=offset)).date()
        counts = activity_map.get(day, {'view': 0, 'download': 0, 'contact': 0})
        daily_analytics.append({
            'day': day,
            'views': counts['view'],
            'downloads': counts['download'],
            'contacts': counts['contact'],
        })

    recent_activities = student.activities.all()[:8]
    nfc_qr_status = 'Live' if student.show_contact_card else 'Hidden'
    public_card_url = reverse('student_contact_card', args=[student.id])

    context = {
        'student': student,
        'display_organization': display_organization,
        'total_engagement': total_engagement,
        'profile_completion': profile_completion,
        'daily_analytics': daily_analytics,
        'recent_activities': recent_activities,
        'nfc_qr_status': nfc_qr_status,
        'public_card_url': public_card_url,
        'contact_template_label': CONTACT_TEMPLATE_META['student_digital_card.html']['label'],
        'print_template_label': _get_print_template_label(student),
    }
    return render(request, 'student_owner_dashboard.html', context)


@login_required
def bulk_upload(request):
    return dashboard_bulk_upload(request)


@login_required
def college_details(request, college_id):
    college = get_object_or_404(College, id=college_id)
    permission_response = _require_school_access(request, college)
    if permission_response:
        return permission_response
    return redirect(f"{reverse('admin_dashboard')}{_build_dashboard_query(college)}")


@login_required
def add_college(request):
    if not _is_super_admin(request.user):
        messages.error(request, 'Only super admins can create schools.')
        return redirect('admin_dashboard')
    if request.method == 'POST':
        college = College(
            name=request.POST.get('name'),
            slogan=request.POST.get('slogan'),
            address=request.POST.get('address'),
            logo=request.FILES.get('logo'),
            principal_name=request.POST.get('principal_name', ''),
            principal_signature=request.FILES.get('principal_signature'),
            website=request.POST.get('website'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            theme_primary=request.POST.get('theme_primary', '#1a3a5c'),
            theme_light_primary=request.POST.get('theme_light_primary', '#f7f5f0'),
            theme_secondary=request.POST.get('theme_secondary', '#1a1a2a'),
            theme_ternary=request.POST.get('theme_ternary', '#c9a84c'),
            description=request.POST.get('description')
        )
        admin_username = request.POST.get('admin_username', '').strip()
        admin_password = request.POST.get('admin_password', '').strip()
        if not admin_username or not admin_password:
            messages.error(request, 'School admin username and password are required.')
            return redirect('admin_dashboard')
        _sync_school_admin_user(college, admin_username, admin_password)
        college.save()
        messages.success(request, "College added successfully!")
        return redirect('admin_dashboard')
    return redirect('admin_dashboard')


@login_required
def edit_college(request, college_id):
    college = get_object_or_404(College, id=college_id)
    permission_response = _require_school_access(request, college)
    if permission_response:
        return permission_response
    if request.method == 'POST':
        return dashboard_settings(request)
    return redirect(f"{reverse('dashboard_settings')}{_build_dashboard_query(college)}")


@login_required
def delete_college(request, college_id):
    college = get_object_or_404(College, id=college_id)
    if not _is_super_admin(request.user):
        messages.error(request, 'Only super admins can delete schools.')
        return redirect('admin_dashboard')
    if request.method == 'POST':
        college.delete()
        messages.success(request, "College deleted successfully!")
        return redirect('admin_dashboard')
    return render(request, 'confirm_delete_college.html', {'college': college})


@login_required
def add_student_to_college(request, college_id):
    college = get_object_or_404(College, id=college_id)
    permission_response = _require_school_access(request, college)
    if permission_response:
        return permission_response
    skills = Skill.objects.all()
    role_type = request.GET.get('role_type', 'student')
    default_role = 'Teacher' if role_type == 'teacher' else 'Student'
    if request.method == 'POST':
        raw_password = _generate_profile_password(request.POST.get('name'))
        member_type = _extract_member_type_from_post(request, role_type)
        student = StudentProfile(
            name=request.POST['name'],
            phone=request.POST['phone'],
            email=request.POST['email'],
            username=request.POST['username'],
            college=college,
            profile_category='school',
            member_type=member_type,
            bio=request.POST.get('bio'),
            role=_extract_role_from_post(request, 'school'),
            address=request.POST.get('address'),
            emergency_contact_name=request.POST.get('emergency_contact_name', ''),
            emergency_contact_phone=request.POST.get('emergency_contact_phone', ''),
            map_url=request.POST.get('map_url') or '',
            academic_level=request.POST.get('academic_level', ''),
            section=request.POST.get('section', ''),
            roll_number=request.POST.get('roll_number', ''),
            blood_group=request.POST.get('blood_group', ''),
            gender=request.POST.get('gender', ''),
            organization_name=request.POST.get('organization_name') or college.name,
            password=raw_password,
            cv=request.FILES.get('cv', None),  # CV upload here
        )
        if request.FILES.get('profile_photo'):
            student.profile_photo = request.FILES['profile_photo']
        if request.FILES.get('cover_photo'):
            student.cover_photo = request.FILES['cover_photo']
        student.portfolio_template = _get_portfolio_template(student) or student.portfolio_template
        student.save()
        _sync_profile_auth_user(student, raw_password)
        student.save(update_fields=['auth_user', 'username'])
        _assign_profile_skills(student, request)
        # Store social links now so they are ready for digital card use later.
        social_fields = ['facebook', 'instagram', 'twitter', 'linkedin', 'youtube', 'tiktok', 'github', 'figma', 'upwork', 'website', 'messenger', 'whatsapp']
        for field in social_fields:
            setattr(student, field, request.POST.get(field))
        student.social_stack = ",".join(
            key for key in ['facebook', 'instagram', 'twitter', 'linkedin', 'youtube', 'tiktok', 'github', 'figma', 'upwork', 'messenger', 'whatsapp']
            if request.POST.get(key)
        )
        student.save()
        if _profile_supports_self_service(student):
            messages.success(request, f'{member_type.title()} added successfully. Default login password: {raw_password}')
        else:
            messages.success(request, f'{member_type.title()} added successfully. This profile is managed by the school admin.')
        return redirect('college_details', college_id=college.id)
    return render(request, 'add_student_to_college.html', {
        'college': college,
        'skills': skills,
        'school_role_choices': SCHOOL_ROLE_CHOICES,
        'student_role_choices': ['Student', 'Class Monitor', 'Head Boy', 'Head Girl', 'Sports Captain'],
        'role_type': role_type,
        'default_role': default_role,
        'academic_level_choices': ACADEMIC_LEVEL_CHOICES,
        'gender_choices': GENDER_CHOICES,
    })
def send_profile_message(request, id):
    student = get_object_or_404(StudentProfile, id=id)
    if request.method == 'POST':
        message = request.POST.get('message')
        send_mail(
            subject='New message via student card',
            message=message,
            from_email=request.user.email,  # or use settings.DEFAULT_FROM_EMAIL
            recipient_list=[student.email],
        )
    return redirect('contact_card', student_id=id)


def student_digital_contact_card(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    if not student.show_contact_card and not _can_manage_profile(request.user, student):
        return render(
            request,
            'contact/student_digital_card.html',
            {'profile_unavailable': True},
            status=403,
        )

    student.views += 1
    student.save(update_fields=['views'])
    _log_profile_activity(student, 'view', 'student-digital-card')
    return render(request, 'contact/student_digital_card.html', _school_card_context(request, student))


def contact_card(request, student_id):
    return student_digital_contact_card(request, student_id)


def download_vcard(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)

    # Increment downloads counter
    student.downloads += 1
    student.save(update_fields=['downloads'])
    _log_profile_activity(student, 'download', 'vcard')

    college_name = student.college.name if student.college else ''
    vcard = f"""BEGIN:VCARD
VERSION:3.0
FN:{student.name}
ORG:{college_name}
TEL;TYPE=CELL:{student.phone}
EMAIL;TYPE=INTERNET:{student.email}
URL:{student.website or ''}
END:VCARD
"""
    response = HttpResponse(vcard, content_type='text/vcard')
    response['Content-Disposition'] = f'attachment; filename=contact_{student.name.replace(" ", "_")}.vcf'
    return response


def track_contact_action(request, student_id, action):
    student = get_object_or_404(StudentProfile, pk=student_id)
    website_url = _normalize_public_url(student.website)

    action_targets = {
        'phone': f'tel:{student.phone}' if student.phone else '',
        'email': f'mailto:{student.email}' if student.email else '',
        'website': website_url,
        'map': _normalize_public_url(student.map_url) or (f'https://www.google.com/maps/search/?api=1&query={student.address.replace(" ", "+")}' if student.address else ''),
        'whatsapp': f'https://wa.me/{student.whatsapp}' if student.whatsapp else '',
        'social-linkedin': student.linkedin or '',
        'social-instagram': student.instagram or '',
        'social-facebook': student.facebook or '',
        'social-messenger': student.messenger or '',
        'social-twitter': student.twitter or '',
        'social-youtube': student.youtube or '',
        'social-tiktok': student.tiktok or '',
        'social-github': student.github or '',
        'social-figma': student.figma or '',
        'social-upwork': student.upwork or '',
    }

    target = action_targets.get(action, '')
    if not target:
        messages.error(request, 'This contact action is not available for this profile.')
        return redirect('contact_card', student_id=student.id)

    student.contact_clicks += 1
    student.save(update_fields=['contact_clicks'])
    _log_profile_activity(student, 'contact', action)
    return redirect(target)


def send_site_message(request):
    if request.method == 'POST':
        contact_type = request.POST.get('contactType')
        full_name = request.POST.get('fullName')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        organization = request.POST.get('organization')
        hear_about = request.POST.get('hearAbout')
        message = request.POST.get('message')
        priority = 'Yes' if request.POST.get('priority') else 'No'

        subject = f"SkillConnect Contact: {contact_type} - {full_name}"
        body = f"""
Contact Form Submission - SkillConnect

Contact Type: {contact_type}
Name: {full_name}
Email: {email}
Phone: {phone}
Organization: {organization}
How they heard about us: {hear_about}
Priority Response: {priority}

Message:
{message}

---
This message was sent from the SkillConnect website contact form.
        """.strip()

        send_mail(
            subject,
            body,
            email,  # from user's email
            [settings.DEFAULT_FROM_EMAIL],  # your Gmail address
        )
        return redirect('home')  # or show a success message

    return redirect('home')


def edit_student_manual(request, student_id):
    student = get_object_or_404(StudentProfile, pk=student_id)
    if not (_can_manage_profile(request.user, student) or _is_student_edit_authorized(request, student.id)):
        messages.error(request, 'Please log in with your profile username and password to edit this card.')
        return redirect('edit_student_auth', student_id=student.id)
    managed_school = _get_managed_school(request.user)
    can_manage_school_fields = _is_super_admin(request.user) or bool(managed_school and managed_school.id == student.college_id)

    contact_templates = CONTACT_TEMPLATES.get(student.user_type, ['student_digital_card.html'])
    portfolio_templates = PORTFOLIO_TEMPLATES.get(student.user_type, [])

    if request.method == 'POST':
        student.name = request.POST.get('name', student.name)
        student.username = request.POST.get('username', student.username)
        student.email = request.POST.get('email', student.email)
        student.phone = request.POST.get('phone', student.phone)
        student.organization_name = request.POST.get('organization_name', student.organization_name)
        student.bio = request.POST.get('bio', student.bio)
        student.role = _extract_role_from_post(request, student.profile_category, student.role)
        student.address = request.POST.get('address', student.address)
        student.emergency_contact_name = request.POST.get('emergency_contact_name', student.emergency_contact_name)
        student.emergency_contact_phone = request.POST.get('emergency_contact_phone', student.emergency_contact_phone)
        student.map_url = request.POST.get('map_url', student.map_url)
        student.blood_group = request.POST.get('blood_group', student.blood_group)
        student.gender = request.POST.get('gender', student.gender)
        student.about_intro = request.POST.get('about_intro', student.about_intro)
        student.about_featured = request.POST.get('about_featured', student.about_featured)
        student.about_current = request.POST.get('about_current', student.about_current)
        student.show_contact_card = request.POST.get('show_contact_card') == 'on'
        student.social_stack = ",".join(request.POST.getlist('social_stack'))
        if can_manage_school_fields:
            student.profile_category = request.POST.get('profile_category', student.profile_category)
            student.member_type = _extract_member_type_from_post(request, student.member_type or 'student')
            student.academic_level = request.POST.get('academic_level', student.academic_level)
            student.section = request.POST.get('section', student.section)
            student.roll_number = request.POST.get('roll_number', student.roll_number)
            if student.profile_category == 'school':
                college_id = request.POST.get('college')
                if not college_id:
                    messages.error(request, 'School / College profiles must use a registered school.')
                    return redirect('edit_student_manual', student_id=student.id)
                student.college = get_object_or_404(College, id=college_id)
            else:
                student.college = None

        if request.FILES.get('profile_photo'):
            student.profile_photo = request.FILES['profile_photo']
        if request.FILES.get('cover_photo'):
            student.cover_photo = request.FILES['cover_photo']
        if request.FILES.get('cv'):
            student.cv = request.FILES['cv']

        student.contact_template = 'student_digital_card.html'
        student.print_card_type = request.POST.get('print_card_type', student.print_card_type or 'id_card')
        student.print_orientation = request.POST.get('print_orientation', student.print_orientation or 'portrait')
        student.print_front_design = request.POST.get('print_front_design', _get_front_design_value(student))
        student.print_back_design = request.POST.get('print_back_design', _get_back_design_value(student))
        student.print_calendar = request.POST.get('print_calendar', student.print_calendar or 'bs')
        student.print_valid_till = request.POST.get('print_valid_till', student.print_valid_till)
        student.print_label = request.POST.get('print_label', student.print_label)
        student.print_template = student.print_front_design
        student.print_custom_note = request.POST.get('print_custom_note', student.print_custom_note)

        for field in ['facebook', 'instagram', 'twitter', 'linkedin', 'youtube', 'tiktok', 'github', 'figma', 'upwork', 'website', 'messenger', 'whatsapp']:
            setattr(student, field, request.POST.get(field))

        student.save()
        _sync_profile_auth_user(student)
        student.save(update_fields=['auth_user', 'username'])
        _assign_profile_skills(student, request)
        messages.success(request, 'Profile settings updated successfully.')
        return redirect('contact_card', student_id=student.id)

    return render(request, 'edit_student_manual.html', {
        'student': student,
        'contact_templates': contact_templates,
        'portfolio_templates': portfolio_templates,
        'colleges': College.objects.all(),
        'can_manage_school_fields': can_manage_school_fields,
        'skills': Skill.objects.all(),
        'selected_socials': [item.strip() for item in student.social_stack.split(',') if item.strip()],
        'social_choices': SOCIAL_CHOICES,  # <-- add this
        'school_role_choices': SCHOOL_ROLE_CHOICES,
        'contact_template_meta': CONTACT_TEMPLATE_META,
        'print_templates': PRINT_TEMPLATES,
        'contact_template_choices': [
            {
                'value': tpl,
                'label': CONTACT_TEMPLATE_META.get(tpl, {}).get('label', tpl),
                'description': CONTACT_TEMPLATE_META.get(tpl, {}).get('description', ''),
            }
            for tpl in contact_templates
        ],
        'print_template_choices': [
            {
                'value': key,
                'label': meta['label'],
                'description': meta['description'],
            }
            for key, meta in PRINT_TEMPLATES.items()
            if key in PRINT_FRONT_THEMES
        ],
        'print_card_types': PRINT_CARD_TYPES,
        'print_orientations': PRINT_ORIENTATIONS,
        'print_front_design_choices': [
            {
                'value': key,
                'label': meta['label'],
                'description': meta['description'],
                'palette': meta['palette'],
            }
            for key, meta in PRINT_FRONT_THEMES.items()
        ],
        'print_back_design_choices': [
            {
                'value': key,
                'label': meta['label'],
                'description': meta['description'],
                'palette': meta['palette'],
            }
            for key, meta in PRINT_BACK_THEMES.items()
        ],
        'academic_level_choices': ACADEMIC_LEVEL_CHOICES,
        'gender_choices': GENDER_CHOICES,
        'member_type_choices': StudentProfile.MEMBER_TYPE_CHOICES,
    })
